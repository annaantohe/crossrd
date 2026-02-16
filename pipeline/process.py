#!/usr/bin/env python3
"""
process.py — main pipeline for crossrd
reads the career framework excel workbook and outputs JSON for the frontend.

usage:
  python process.py --family healthcare
  python process.py --input ../data/healthcare/career_framework_v4.xlsx --output ../src/data/healthcare.json
  python process.py --validate ../src/data/healthcare.json
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

import openpyxl

from config import (
    CATEGORIES, RADAR_DIMENSIONS, RADAR_CATEGORY_MAP,
    load_family_config, list_families,
)
from financial import read_net_worth_trajectory, read_financial_assumptions, build_money_data, build_timeline_data
from stress import read_stress_test


def read_career_tracks(wb):
    """read the master list of career tracks from the Career Tracks sheet"""
    ws = wb["Career Tracks"]
    tracks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        tracks.append({
            "id": int(row[0]),
            "profession": row[1],
            "name": row[2],
            "category": row[3],
            "procedural_intensity": row[4],
            "status": row[5],
        })
    print(f"  read {len(tracks)} career tracks")
    return tracks


def read_l2_matrix(wb, profession, l2_sheets, l2_columns, finalists):
    """read the raw specialty data from an L2 Matrix sheet.

    args:
        wb: openpyxl workbook
        profession: profession key (e.g. "MD/DO")
        l2_sheets: dict mapping profession -> Excel sheet name
        l2_columns: dict mapping column index -> field name
        finalists: dict mapping track name -> finalist info
    """
    sheet_name = l2_sheets[profession]
    ws = wb[sheet_name]

    specialties = []
    # data starts at row 5 (row 1 = title, row 2 = category tags, row 3 = headers, row 4 = Dec/Ref)
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None:
            break

        spec = {"profession": profession}
        for col_idx, field_name in l2_columns.items():
            val = row[col_idx]
            # convert numeric strings to numbers
            if field_name != "name" and val is not None:
                try:
                    val = float(val) if isinstance(val, (int, float)) else float(val)
                    # keep as int if it's a whole number
                    if val == int(val):
                        val = int(val)
                except (ValueError, TypeError):
                    pass
            spec[field_name] = val

        # check if this track is a finalist
        spec["finalist"] = spec["name"] in finalists
        if spec["finalist"]:
            info = finalists[spec["name"]]
            spec["finalist_key"] = info["key"]
            spec["teen_name"] = info["teen_name"]
            spec["finalist_color"] = info["color"]

        specialties.append(spec)

    print(f"  read {len(specialties)} specialties from {sheet_name}")
    return specialties


def read_finals_matrix(wb, finalist_order, scenario_map):
    """read the category scores and scenario totals for the finalists.

    args:
        wb: openpyxl workbook
        finalist_order: list of finalist keys in column order
        scenario_map: dict mapping Excel display names to slug keys
    """
    ws = wb["Finals Matrix"]

    # read category scores (rows 4-17, cols D-I = indices 3-8)
    category_scores = {}  # key -> {cat_id: score}

    for row in ws.iter_rows(min_row=4, max_row=17, values_only=True):
        if row[0] is None:
            continue
        # skip the header row with "#"
        try:
            cat_id = int(row[0])
        except (ValueError, TypeError):
            continue
        for i, key in enumerate(finalist_order):
            if key not in category_scores:
                category_scores[key] = {}
            val = row[3 + i]
            if val is not None:
                category_scores[key][cat_id] = round(float(val), 2)

    # read scenario totals (starts after "SCENARIO-WEIGHTED TOTAL SCORES" header)
    scenario_totals = {}  # key -> {scenario: score}
    scenario_rows = []
    found_scenarios = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] == "Scenario":
            found_scenarios = True
            continue
        if found_scenarios and row[0] is not None:
            display_name = row[0]
            if display_name.startswith("KEY"):
                break
            scores = {}
            for i, key in enumerate(finalist_order):
                val = row[1 + i]
                if val is not None:
                    scores[key] = round(float(val), 2)
            scenario_rows.append((display_name, scores))

    for display_name, scores in scenario_rows:
        scenario_key = scenario_map.get(display_name, display_name.lower().replace(" ", "_"))
        for finalist_key, score in scores.items():
            if finalist_key not in scenario_totals:
                scenario_totals[finalist_key] = {}
            scenario_totals[finalist_key][scenario_key] = score

    print(f"  read finals matrix: {len(category_scores)} finalists, {len(scenario_rows)} scenarios")
    return category_scores, scenario_totals


def read_scenario_profiles(wb):
    """read the scenario weight profiles from the Scenario Profiles sheet"""
    ws = wb["Scenario Profiles"]

    # column mapping: B=Equal Weight, C=Max Earnings, D=Best Lifestyle,
    # E=Fastest to Practice, F=Most Procedural
    profiles = {"default": {}}
    scenario_cols = {
        1: "equal_weight",
        2: "max_earnings",
        3: "best_lifestyle",
        4: "fastest_to_practice",
        5: "most_procedural",
    }

    for cat_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=15, values_only=True)):
        cat_id = cat_idx + 1
        # default weights come from universal categories
        profiles["default"][str(cat_id)] = CATEGORIES[cat_idx]["weight"]

        for col_offset, scenario_name in scenario_cols.items():
            if scenario_name not in profiles:
                profiles[scenario_name] = {}
            val = row[col_offset]
            profiles[scenario_name][str(cat_id)] = int(val) if val is not None else 7

    return profiles


def read_radar_data(finals_scores, finalist_order):
    """build the 6-dimension radar chart data from the finals category scores.

    uses the RADAR_CATEGORY_MAP to consolidate 14 categories into 6 dimensions.
    each dimension is the average of its mapped category scores.
    """
    radar = []
    for dim_info in RADAR_DIMENSIONS:
        point = {"dim": dim_info["dim"], "emoji": dim_info["emoji"]}
        cat_ids = RADAR_CATEGORY_MAP[dim_info["dim"]]

        for key in finalist_order:
            scores = finals_scores.get(key, {})
            vals = [scores.get(c, 5) for c in cat_ids]
            point[key] = round(sum(vals) / len(vals), 1)

        radar.append(point)
    return radar


def build_tracks_json(all_specialties, finals_scores, scenario_totals):
    """build the tracks array for the output json"""
    tracks = []
    for spec in all_specialties:
        track = {
            "name": spec["name"],
            "profession": spec["profession"],
            "finalist": spec.get("finalist", False),
            "raw_data": {
                "startSalary": spec.get("startSalary", 0),
                "midSalary": spec.get("midSalary", 0),
                "peakSalary": spec.get("peakSalary", 0),
                "hoursWeek": spec.get("hoursWeek", 0),
                "burnout": spec.get("burnout", 0),
                "satisfaction": spec.get("satisfaction", 0),
                "chooseAgain": spec.get("chooseAgain", 0),
                "malpracticeCost": spec.get("malpracticeCost", 0),
                "vacation": spec.get("vacation", 0),
                "matchComp": spec.get("matchComp", 0),
                "callSchedule": spec.get("callSchedule", 0),
                "physicalToll": spec.get("physicalToll", 0),
                "emotionalToll": spec.get("emotionalToll", 0),
            },
        }

        if spec.get("finalist"):
            key = spec["finalist_key"]
            track["teen_name"] = spec["teen_name"]
            track["finalist_color"] = spec["finalist_color"]
            track["finalist_key"] = key
            track["scores"] = {
                f"category_{cat_id}": score
                for cat_id, score in finals_scores.get(key, {}).items()
            }
            track["scenario_totals"] = scenario_totals.get(key, {})

        tracks.append(track)

    return tracks


def build_careers_array(cfg):
    """build the careers array for the frontend from the family config.

    this replaces the hardcoded CAREERS constant in theme.js — the frontend
    reads this from the JSON instead.
    """
    careers = []
    for name, info in cfg["finalists"].items():
        # figure out which profession this finalist belongs to
        path = ""
        for prof_key, prof_info in cfg["professions"].items():
            path = prof_info["label"]
            # we don't have a direct mapping, so we'll set path later
            # when we match against the track data
        careers.append({
            "key": info["key"],
            "name": info["teen_name"],
            "color": info["color"],
        })
    return careers


def process(family_slug, input_path=None, output_path=None):
    """main pipeline: read excel, build json, write output.

    args:
        family_slug: the profession family slug (e.g. "healthcare")
        input_path: override path to excel workbook (optional)
        output_path: override path for output json (optional)
    """
    repo_root = Path(__file__).parent.parent
    cfg = load_family_config(family_slug)

    if input_path is None:
        input_path = repo_root / "data" / family_slug / cfg["source"]
    if output_path is None:
        output_path = repo_root / "src" / "data" / f"{family_slug}.json"

    # pull config values
    finalists = cfg["finalists"]
    finalist_order = cfg["finalist_order"]
    l2_sheets = cfg["l2_sheets"]
    l2_columns = {int(k): v for k, v in cfg["l2_columns"].items()}
    stress_key_map = cfg.get("stress_key_map", {})
    scenario_map = cfg.get("scenario_map", {})
    final_ranking = cfg.get("final_ranking", [])
    decision_tree = cfg.get("decision_tree", {})
    decision_tree_results = cfg.get("decision_tree_results", {})

    print(f"processing {family_slug}...")
    print(f"  reading {input_path}...")
    wb = openpyxl.load_workbook(str(input_path), read_only=True, data_only=True)

    # 1. read all the raw data
    print("reading career tracks...")
    career_tracks = read_career_tracks(wb)

    print("reading L2 matrix data...")
    all_specialties = []
    for profession in l2_sheets:
        specs = read_l2_matrix(wb, profession, l2_sheets, l2_columns, finalists)
        all_specialties.extend(specs)

    print("reading finals matrix...")
    finals_scores, scenario_totals = read_finals_matrix(wb, finalist_order, scenario_map)

    print("reading scenario profiles...")
    scenario_profiles = read_scenario_profiles(wb)

    print("reading financial data...")
    net_worth = read_net_worth_trajectory(wb, finalist_order)
    money_data = build_money_data(wb, finalist_order)

    print("reading stress test...")
    stress_data = read_stress_test(wb, stress_key_map)

    wb.close()

    # 2. build derived data
    print("building radar chart data...")
    radar_data = read_radar_data(finals_scores, finalist_order)

    print("building timeline data...")
    timeline_data = build_timeline_data(cfg)

    print("building tracks json...")
    tracks = build_tracks_json(all_specialties, finals_scores, scenario_totals)

    # build careers array — add the profession path by looking at the track data
    careers = []
    for name, info in finalists.items():
        # find this finalist in the track data to get their profession
        track_match = next((s for s in all_specialties if s["name"] == name), None)
        prof_key = track_match["profession"] if track_match else ""
        prof_label = cfg["professions"].get(prof_key, {}).get("label", prof_key)
        careers.append({
            "key": info["key"],
            "name": info["teen_name"],
            "color": info["color"],
            "path": prof_label,
        })

    # 3. assemble the output
    output = {
        "meta": {
            "profession_family": cfg["slug"],
            "family_name": cfg["name"],
            "headline": cfg.get("headline", f"Career Guide: {cfg['name']}"),
            "subtitle": cfg.get("subtitle", "A Data-Driven Guide"),
            "icon": cfg.get("icon", ""),
            "last_updated": str(date.today()),
            "total_tracks": len(all_specialties),
            "finalists": sum(1 for t in tracks if t["finalist"]),
            "data_points": 107,
            "source_file": cfg["source"],
        },
        "professions": {
            prof: {
                "label": info["label"],
                "color": info["color"],
                "track_count": sum(1 for s in all_specialties if s["profession"] == prof),
            }
            for prof, info in cfg["professions"].items()
        },
        "careers": careers,
        "categories": [
            {
                "id": cat["id"],
                "name": cat["name"],
                "weight_default": cat["weight"],
                "description": cat["description"],
            }
            for cat in CATEGORIES
        ],
        "scenario_profiles": scenario_profiles,
        "tracks": tracks,
        "finalists": {
            "ranking": final_ranking,
            "net_worth_trajectory": net_worth,
            "radar_dimensions": radar_data,
            "stress_test": stress_data,
            "money_data": money_data,
            "timeline_data": timeline_data,
            "decision_tree": decision_tree,
            "decision_tree_results": decision_tree_results,
        },
    }

    # 4. write it out
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with open(output_file, "w") as f:
        json.dump(output, f, indent=2)

    print(f"\ndone! wrote {output_file}")
    print(f"  {output['meta']['total_tracks']} tracks, {output['meta']['finalists']} finalists")
    return output


def validate(json_path):
    """validate the output json against expected values from the family config"""
    print(f"validating {json_path}...")
    with open(json_path) as f:
        data = json.load(f)

    errors = []

    family = data["meta"]["profession_family"]
    try:
        cfg = load_family_config(family)
    except FileNotFoundError:
        print(f"  warning: no config found for '{family}', using basic validation")
        cfg = None

    # check finalist count
    if cfg:
        expected_finalists = len(cfg["finalists"])
        if data["meta"]["finalists"] != expected_finalists:
            errors.append(f"expected {expected_finalists} finalists, got {data['meta']['finalists']}")

        # check all finalist keys are present
        for name, info in cfg["finalists"].items():
            key = info["key"]
            found = any(t.get("finalist_key") == key for t in data["tracks"])
            if not found:
                errors.append(f"finalist '{key}' ({name}) not found in tracks")
    else:
        # fallback: basic checks
        if data["meta"]["finalists"] < 1:
            errors.append("no finalists found")

    # check we have tracks
    if data["meta"]["total_tracks"] < 1:
        errors.append("no tracks found")

    # check net worth trajectory exists
    if len(data["finalists"]["net_worth_trajectory"]) < 1:
        errors.append("no net worth trajectory data")

    # check stress test exists
    if len(data["finalists"]["stress_test"]) < 1:
        errors.append("no stress test data")

    # check careers array exists
    if "careers" not in data or len(data["careers"]) < 1:
        errors.append("no careers array in output")

    if errors:
        print(f"\nFAILED — {len(errors)} errors:")
        for e in errors:
            print(f"  ✗ {e}")
        return False
    else:
        print("\nPASSED — all checks ok ✓")
        return True


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="crossrd data pipeline")
    parser.add_argument("--family", help="profession family slug (e.g. healthcare)")
    parser.add_argument("--all", action="store_true", help="process all registered families")
    parser.add_argument("--input", help="path to the excel workbook (overrides family config)")
    parser.add_argument("--output", help="path for the output json (overrides family config)")
    parser.add_argument("--validate", help="validate an existing json file")
    args = parser.parse_args()

    if args.validate:
        ok = validate(args.validate)
        sys.exit(0 if ok else 1)
    elif args.all:
        families = list_families()
        if not families:
            print("no families found (no data/*/config.yaml files)")
            sys.exit(1)
        print(f"processing {len(families)} families: {', '.join(families)}")
        for fam in families:
            process(fam)
    elif args.family:
        process(args.family, args.input, args.output)
    elif args.input and args.output:
        # backwards compat: if --input and --output given without --family,
        # try to guess the family from the input path
        input_path = Path(args.input)
        family_slug = input_path.parent.name
        process(family_slug, args.input, args.output)
    else:
        parser.print_help()
        sys.exit(1)

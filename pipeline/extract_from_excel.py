#!/usr/bin/env python3
"""
extract_from_excel.py — one-time migration tool

reads the career framework excel workbook and writes YAML data files
that replace the excel as the pipeline's source of truth.

usage:
    python extract_from_excel.py --family healthcare
"""

import argparse
import sys
from pathlib import Path

import openpyxl
import yaml

from config import CATEGORIES, load_family_config


def extract_specialties(wb, l2_sheets, l2_columns):
    """extract raw specialty data from L2 Matrix sheets.

    returns dict: {profession: [list of specialty dicts]}
    """
    result = {}
    for profession, sheet_name in l2_sheets.items():
        ws = wb[sheet_name]
        specialties = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] is None:
                break
            spec = {}
            for col_idx, field_name in l2_columns.items():
                val = row[col_idx]
                if field_name != "name" and val is not None:
                    try:
                        val = float(val)
                        if val == int(val):
                            val = int(val)
                    except (ValueError, TypeError):
                        pass
                spec[field_name] = val
            specialties.append(spec)
        result[profession] = specialties
        print(f"  extracted {len(specialties)} specialties from {sheet_name}")
    return result


def extract_l1_scores(wb):
    """extract L1 scoring data from L1 Scoring Results sheet.

    returns two dicts:
      data_points: {profession: {cat_id: [list of scores]}}
      averages: {profession: {cat_id: float}}
    """
    ws = wb["L1 Scoring Results"]
    professions = ["MD/DO", "DDS/DMD", "DPM", "OD"]
    prof_col = {"MD/DO": 2, "DDS/DMD": 3, "DPM": 4, "OD": 5}

    data_points = {p: {} for p in professions}
    current_cat = None

    for row in ws.iter_rows(min_row=3, max_row=55, values_only=True):
        cell0 = row[0]
        cell1 = str(row[1]) if row[1] else ""

        # detect category headers like "Cat 1: Pre-Professional Phase"
        if cell0 is None and "Cat " in cell1:
            # extract category number
            cat_num = int(cell1.split("Cat ")[1].split(":")[0])
            current_cat = cat_num
            continue

        # data point row: col 0 has a number (data point ID)
        if cell0 is not None and isinstance(cell0, (int, float)) and current_cat is not None:
            for prof in professions:
                val = row[prof_col[prof]]
                if val is not None:
                    try:
                        score = float(val)
                        if score == int(score):
                            score = int(score)
                    except (ValueError, TypeError):
                        continue
                    if current_cat not in data_points[prof]:
                        data_points[prof][current_cat] = []
                    data_points[prof][current_cat].append(score)

        # proxy score rows (Cat 8 and Cat 9 have proxy scores)
        if cell0 is None and "Profession-level proxy score" in cell1:
            for prof in professions:
                val = row[prof_col[prof]]
                if val is not None:
                    try:
                        score = float(val)
                        if score == int(score):
                            score = int(score)
                    except (ValueError, TypeError):
                        continue
                    if current_cat not in data_points[prof]:
                        data_points[prof][current_cat] = []
                    data_points[prof][current_cat].append(score)

    # also read the category average scores (rows 60-73)
    averages = {p: {} for p in professions}
    for row in ws.iter_rows(min_row=60, max_row=73, values_only=True):
        if row[0] is None:
            continue
        try:
            cat_id = int(row[0])
        except (ValueError, TypeError):
            continue
        for prof in professions:
            val = row[prof_col[prof]]
            if val is not None:
                averages[prof][cat_id] = round(float(val), 2)

    print(f"  extracted L1 data points for {len(professions)} professions")
    return data_points, averages


def extract_scenario_profiles(wb):
    """extract scenario weight profiles from Scenario Profiles sheet."""
    ws = wb["Scenario Profiles"]

    profiles = {"default": {}}
    scenario_cols = {
        1: "equal_weight",
        2: "max_earnings",
        3: "best_lifestyle",
        4: "fastest_to_practice",
        5: "most_procedural",
    }

    for cat_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=15, values_only=True)):
        cat_id = cat_idx + 1
        profiles["default"][cat_id] = CATEGORIES[cat_idx]["weight"]

        for col_offset, scenario_name in scenario_cols.items():
            if scenario_name not in profiles:
                profiles[scenario_name] = {}
            val = row[col_offset]
            profiles[scenario_name][cat_id] = int(val) if val is not None else 7

    print(f"  extracted {len(profiles)} scenario profiles")
    return profiles


def extract_financial_model(wb, finalist_order):
    """extract financial assumptions from Career Life Models sheet."""
    ws = wb["Career Life Models"]

    # map row labels to our YAML keys (must match exact Excel labels)
    label_map = {
        "Undergrad Cost ($K/yr)": "undergrad_cost_per_yr",
        "Undergrad Years": "undergrad_years",
        "Prof School Cost ($K/yr)": "prof_school_cost_per_yr",
        "Prof School Years": "prof_school_years",
        "Residency Years": "residency_years",
        "Fellowship Years": "fellowship_years",
        "Trainee Salary ($K/yr)": "trainee_salary",
        "Age at Independent Practice": "age_independent",
        "Starting Salary ($K)": "starting_salary",
        "Mid-Career Salary ($K, age 40)": "mid_salary",
        "Peak Salary ($K, age 48+)": "peak_salary",
        "Malpractice ($K/yr)": "malpractice_per_yr",
        "Annual Overhead/CE ($K/yr)": "overhead_per_yr",
        "Total Education Debt ($K)": "education_debt",
        "Student Loan Rate (%)": "loan_rate",
        "Living Expenses ($K/yr)": "living_expenses",
        "Living Exp Growth (%/yr)": "living_exp_growth",
        "Salary Growth to Mid (%/yr)": "salary_growth_to_mid",
        "Salary Growth to Peak (%/yr)": "salary_growth_to_peak",
        "Post-Peak Maintenance (%/yr)": "post_peak_growth",
        "NPV Discount Rate (%)": "npv_discount_rate",
    }

    finalists = {key: {} for key in finalist_order}

    for row in ws.iter_rows(min_row=5, max_row=38, values_only=True):
        if row[0] is None:
            continue
        label = str(row[0]).strip()
        yaml_key = label_map.get(label)
        if yaml_key is None:
            continue
        for i, key in enumerate(finalist_order):
            val = row[1 + i]
            if val is not None:
                try:
                    val = float(val)
                    if val == int(val):
                        val = int(val)
                except (ValueError, TypeError):
                    pass
                finalists[key][yaml_key] = val

    # also extract the net worth trajectory
    trajectory = []
    found_header = False
    for row in ws.iter_rows(min_row=39, values_only=True):
        if row[0] == "Age":
            found_header = True
            continue
        if found_header:
            if row[0] is None or (isinstance(row[0], str) and not row[0].isdigit()):
                break
            age = int(row[0])
            point = {"age": age}
            for i, key in enumerate(finalist_order):
                val = row[1 + i]
                point[key] = int(val) if val is not None else 0
            trajectory.append(point)

    print(f"  extracted financial model for {len(finalists)} finalists, {len(trajectory)} trajectory points")
    return finalists, trajectory


def extract_stress_test(wb, key_map):
    """extract stress test scores from Stress Test sheet."""
    ws = wb["Stress Test"]

    stress_data = {}
    found_composite = False
    found_header = False

    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] == "COMPOSITE STRESS TEST RESILIENCE":
            found_composite = True
            continue
        if found_composite and row[0] == "Track":
            found_header = True
            continue
        if found_header:
            if row[0] is None or row[0] == "STRESS TEST INSIGHT":
                break
            track_name = row[0]
            key = key_map.get(track_name, track_name.lower())
            stress_data[key] = {
                "ai": int(row[1]) if row[1] else 0,
                "pay": int(row[2]) if row[2] else 0,
                "injury": int(row[3]) if row[3] else 0,
                "match": int(row[4]) if row[4] else 0,
            }

    print(f"  extracted stress test for {len(stress_data)} finalists")
    return stress_data


def extract_finals_matrix(wb, finalist_order, scenario_map):
    """extract the pre-computed category scores and scenario totals from Finals Matrix.

    these serve as the "ground truth" for validating the scoring engine.
    """
    ws = wb["Finals Matrix"]

    category_scores = {key: {} for key in finalist_order}
    for row in ws.iter_rows(min_row=4, max_row=18, values_only=True):
        if row[0] is None:
            continue
        try:
            cat_id = int(row[0])
        except (ValueError, TypeError):
            continue
        for i, key in enumerate(finalist_order):
            val = row[3 + i]
            if val is not None:
                category_scores[key][cat_id] = round(float(val), 2)

    scenario_totals = {key: {} for key in finalist_order}
    found_scenarios = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] == "Scenario":
            found_scenarios = True
            continue
        if found_scenarios and row[0] is not None:
            display_name = str(row[0])
            if display_name.startswith("KEY"):
                break
            scenario_key = scenario_map.get(display_name, display_name.lower().replace(" ", "_"))
            for i, key in enumerate(finalist_order):
                val = row[1 + i]
                if val is not None:
                    scenario_totals[key][scenario_key] = round(float(val), 2)

    print(f"  extracted finals matrix: {len(finalist_order)} finalists, {len(category_scores[finalist_order[0]])} categories")
    return category_scores, scenario_totals


def write_yaml(data, path, comment=None):
    """write data to a YAML file with optional header comment."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        if comment:
            f.write(comment + "\n")
        yaml.dump(data, f, default_flow_style=False, sort_keys=False, allow_unicode=True)
    print(f"  wrote {path}")


def main():
    parser = argparse.ArgumentParser(description="Extract Excel data to YAML files")
    parser.add_argument("--family", required=True, help="profession family slug")
    args = parser.parse_args()

    repo_root = Path(__file__).parent.parent
    cfg = load_family_config(args.family)
    data_dir = repo_root / "data" / args.family

    input_path = data_dir / cfg["source"]
    print(f"reading {input_path}...")
    wb = openpyxl.load_workbook(str(input_path), read_only=True, data_only=True)

    l2_columns = {int(k): v for k, v in cfg["l2_columns"].items()}
    finalist_order = cfg["finalist_order"]
    stress_key_map = cfg.get("stress_key_map", {})
    scenario_map = cfg.get("scenario_map", {})
    finalists_cfg = cfg["finalists"]

    # 1. extract specialty raw data
    print("\nextracting L2 specialty data...")
    specialties_by_prof = extract_specialties(wb, cfg["l2_sheets"], l2_columns)

    for prof, specs in specialties_by_prof.items():
        slug = prof.lower().replace("/", "_")
        yaml_data = {"profession": prof, "specialties": specs}
        write_yaml(
            yaml_data,
            data_dir / "specialties" / f"{slug}.yaml",
            f"# {prof} specialty raw data — extracted from {cfg['source']}",
        )

    # 2. extract L1 scores
    print("\nextracting L1 scores...")
    l1_data_points, l1_averages = extract_l1_scores(wb)

    l1_yaml = {"professions": {}}
    for prof in l1_data_points:
        prof_data = {"l1_only": {}, "l1_data_points": {}}

        # determine which categories are L1-only vs mixed
        # L1-only: cats where all finalists in this profession have identical finals scores
        # We'll store data points either way, and also store the computed averages
        for cat_id, scores in l1_data_points[prof].items():
            avg = l1_averages.get(prof, {}).get(cat_id)
            prof_data["l1_data_points"][cat_id] = {
                "scores": scores,
                "count": len(scores),
                "average": avg if avg is not None else round(sum(scores) / len(scores), 2),
            }

        l1_yaml["professions"][prof] = prof_data

    write_yaml(
        l1_yaml,
        data_dir / "l1_scores.yaml",
        "# L1 profession-level category scores — extracted from L1 Scoring Results sheet",
    )

    # 3. extract scenario profiles
    print("\nextracting scenario profiles...")
    profiles = extract_scenario_profiles(wb)
    write_yaml(
        {"profiles": profiles},
        data_dir / "scenario_profiles.yaml",
        "# scenario weight profiles — extracted from Scenario Profiles sheet",
    )

    # 4. extract financial model
    print("\nextracting financial model...")
    fin_params, trajectory = extract_financial_model(wb, finalist_order)
    fin_yaml = {
        "finalists": fin_params,
        "reference_trajectory": trajectory,
    }
    write_yaml(
        fin_yaml,
        data_dir / "financial_model.yaml",
        "# financial model assumptions and reference trajectory — extracted from Career Life Models",
    )

    # 5. extract stress test
    print("\nextracting stress test...")
    stress = extract_stress_test(wb, stress_key_map)
    stress_yaml = {
        "scenarios": [
            {"id": "ai", "name": "AI Disruption Accelerates"},
            {"id": "pay", "name": "Reimbursement Cuts 20%"},
            {"id": "injury", "name": "Career-Ending Injury at 40"},
            {"id": "match", "name": "Failure to Match"},
        ],
        "scores": stress,
    }
    write_yaml(
        stress_yaml,
        data_dir / "stress_test.yaml",
        "# stress test resilience scores — extracted from Stress Test sheet",
    )

    # 6. extract finals matrix (ground truth for validation)
    print("\nextracting finals matrix (ground truth)...")
    cat_scores, scen_totals = extract_finals_matrix(wb, finalist_order, scenario_map)
    ground_truth = {
        "category_scores": cat_scores,
        "scenario_totals": scen_totals,
    }
    write_yaml(
        ground_truth,
        data_dir / "ground_truth.yaml",
        "# pre-computed category scores and scenario totals from Excel — for validation only",
    )

    wb.close()
    print(f"\ndone! all YAML files written to {data_dir}/")


if __name__ == "__main__":
    main()

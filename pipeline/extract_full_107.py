#!/usr/bin/env python3
"""
extract_full_107.py — extract all 107 data points from the Excel workbook

Phase 1 of the full-107 migration:
1. Extracts the Data Points Framework (all 107 data point definitions)
2. Extracts the 11 missing L2 columns from the L2 Matrix sheets
3. Enriches l1_scores.yaml with labels and data point IDs

Usage:
    python extract_full_107.py --family healthcare
"""

import argparse
import re
from pathlib import Path

import openpyxl
import yaml

from config import load_family_config


def extract_data_points_framework(wb):
    """Extract all 107 data point definitions from the Data Points Framework sheet."""
    ws = wb["Data Points Framework"]
    data_points = []
    for row in ws.iter_rows(min_row=2, max_row=200, values_only=True):
        if row[0] is None:
            break
        cat_str = str(row[1])
        # parse category id from "1. Pre-Professional Phase" etc.
        m = re.match(r"(\d+)\.", cat_str)
        cat_id = int(m.group(1)) if m else 0
        cat_name = cat_str.split(". ", 1)[-1] if ". " in cat_str else cat_str

        data_points.append({
            "id": int(row[0]),
            "category_id": cat_id,
            "category_name": cat_name,
            "name": str(row[2]).strip(),
            "layer": str(row[3]).strip(),
            "data_type": str(row[4]).strip(),
            "decision_weight": str(row[5]).strip(),
            "scoring_notes": str(row[6]).strip() if row[6] else None,
        })
    return data_points


# columns missing from the original extraction (cols 2, 7, 8, 10, 11, 15, 24, 25, 33)
MISSING_L2_COLUMNS = {
    2: "fellowshipOptions",
    7: "compModel",
    8: "revPerPatientVisit",
    10: "dailyWorkflow",
    11: "patientVolume",
    15: "supportStaffModel",
    24: "geographicDemandNotes",
    25: "telehealthPotential",
    33: "genderDistribution",
}


def extract_missing_l2_columns(wb, l2_sheets):
    """Extract the 9 L2 columns that were skipped in the original extraction.

    Returns dict: {profession: {specialty_name: {field: value}}}
    """
    result = {}
    for profession, sheet_name in l2_sheets.items():
        ws = wb[sheet_name]
        result[profession] = {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] is None:
                break
            name = str(row[0]).strip()
            fields = {}
            for col_idx, field_name in MISSING_L2_COLUMNS.items():
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    val = str(val).strip() if isinstance(val, str) else val
                    # try to convert to number for numeric fields
                    if field_name == "telehealthPotential" and val is not None:
                        try:
                            val = float(val)
                            if val == int(val):
                                val = int(val)
                        except (ValueError, TypeError):
                            pass
                fields[field_name] = val if val else None
            result[profession][name] = fields
        print(f"  extracted {len(result[profession])} specialties' missing columns from {sheet_name}")
    return result


def extract_l1_labels(wb):
    """Extract data point labels from L1 Scoring Results to enrich l1_scores.yaml.

    Returns dict: {category_id: [(data_point_id, label, {prof: score})]}
    """
    ws = wb["L1 Scoring Results"]
    prof_col = {"MD/DO": 2, "DDS/DMD": 3, "DPM": 4, "OD": 5}
    professions = list(prof_col.keys())

    categories = {}
    current_cat = None

    for row in ws.iter_rows(min_row=3, max_row=55, values_only=True):
        cell0 = row[0]
        cell1 = str(row[1]) if row[1] else ""

        # detect category headers
        if cell0 is None and "Cat " in cell1:
            m = re.search(r"Cat (\d+)", cell1)
            if m:
                current_cat = int(m.group(1))
                if current_cat not in categories:
                    categories[current_cat] = []
            continue

        # data point row
        if cell0 is not None and current_cat is not None:
            try:
                dp_id = int(float(cell0))
            except (ValueError, TypeError):
                continue
            label = cell1.strip() if cell1 else f"Data Point {dp_id}"
            scores = {}
            for prof in professions:
                val = row[prof_col[prof]]
                if val is not None:
                    try:
                        scores[prof] = int(float(val))
                    except (ValueError, TypeError):
                        pass
            categories[current_cat].append({
                "data_point_id": dp_id,
                "label": label,
                "scores": scores,
            })

        # proxy scores for Cat 8, 9
        if cell0 is None and "Profession-level proxy score" in cell1 and current_cat:
            scores = {}
            for prof in professions:
                val = row[prof_col[prof]]
                if val is not None:
                    try:
                        scores[prof] = int(float(val))
                    except (ValueError, TypeError):
                        pass
            categories[current_cat].append({
                "data_point_id": None,
                "label": "Profession-level proxy score",
                "scores": scores,
            })

    return categories


def merge_missing_columns_into_yaml(data_dir, missing_data):
    """Patch existing specialties/*.yaml files with the missing L2 columns."""
    for profession, specs_data in missing_data.items():
        slug = profession.lower().replace("/", "_")
        yaml_path = data_dir / "specialties" / f"{slug}.yaml"

        if not yaml_path.exists():
            print(f"  warning: {yaml_path} not found, skipping")
            continue

        with open(yaml_path) as f:
            existing = yaml.safe_load(f)

        updated = 0
        for spec in existing.get("specialties", []):
            name = spec.get("name", "")
            if name in specs_data:
                for field, val in specs_data[name].items():
                    spec[field] = val
                updated += 1

        with open(yaml_path, "w") as f:
            # keep the header comment
            f.write(f"# {profession} specialty raw data — full 38 columns\n")
            yaml.dump(existing, f, default_flow_style=False, sort_keys=False, allow_unicode=True)

        print(f"  patched {updated} specialties in {yaml_path.name}")


def build_enriched_l1_scores(l1_labels, data_dir):
    """Build enriched l1_scores.yaml with labels and data_point_ids."""
    yaml_path = data_dir / "l1_scores.yaml"
    with open(yaml_path) as f:
        existing = yaml.safe_load(f)

    professions = ["MD/DO", "DDS/DMD", "DPM", "OD"]

    for prof in professions:
        prof_data = existing.get("professions", {}).get(prof, {})
        l1_dp = prof_data.get("l1_data_points", {})

        for cat_id, entries in l1_labels.items():
            cat_data = l1_dp.get(cat_id)
            if cat_data is None:
                continue

            labels = []
            dp_ids = []
            for entry in entries:
                if prof in entry["scores"]:
                    labels.append(entry["label"])
                    dp_ids.append(entry["data_point_id"])

            if labels:
                cat_data["labels"] = labels
                cat_data["data_point_ids"] = dp_ids

    with open(yaml_path, "w") as f:
        f.write("# L1 profession-level category scores — enriched with data point labels\n")
        yaml.dump(existing, f, default_flow_style=False, sort_keys=False, allow_unicode=True)

    print(f"  enriched {yaml_path.name} with labels")


def write_yaml(data, path, comment=None):
    """Write data to a YAML file with optional header comment."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        if comment:
            f.write(comment + "\n")
        yaml.dump(data, f, default_flow_style=False, sort_keys=False, allow_unicode=True)
    print(f"  wrote {path}")


def main():
    parser = argparse.ArgumentParser(description="Extract all 107 data points from Excel")
    parser.add_argument("--family", required=True, help="profession family slug")
    args = parser.parse_args()

    repo_root = Path(__file__).parent.parent
    data_dir = repo_root / "data" / args.family

    excel_path = data_dir / "archive" / "career_framework_v4.xlsx"
    if not excel_path.exists():
        print(f"error: {excel_path} not found")
        return

    print(f"reading {excel_path}...")
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    # 1. Extract Data Points Framework
    print("\n1. Extracting Data Points Framework...")
    dp_framework = extract_data_points_framework(wb)
    write_yaml(
        {"data_points": dp_framework},
        data_dir / "data_points_framework.yaml",
        "# All 107 data points — master catalog extracted from Excel Data Points Framework sheet",
    )
    print(f"  {len(dp_framework)} data points extracted")

    # 2. Extract missing L2 columns
    print("\n2. Extracting missing L2 columns...")
    l2_sheets = {
        "MD/DO": "L2 Matrix — MD-DO",
        "DDS/DMD": "L2 Matrix — DDS-DMD",
        "DPM": "L2 Matrix — DPM",
        "OD": "L2 Matrix — OD",
    }
    missing_data = extract_missing_l2_columns(wb, l2_sheets)
    merge_missing_columns_into_yaml(data_dir, missing_data)

    # 3. Extract L1 labels
    print("\n3. Enriching L1 scores with labels...")
    l1_labels = extract_l1_labels(wb)
    build_enriched_l1_scores(l1_labels, data_dir)

    wb.close()
    print(f"\ndone! all 107 data points extracted to {data_dir}/")


if __name__ == "__main__":
    main()
